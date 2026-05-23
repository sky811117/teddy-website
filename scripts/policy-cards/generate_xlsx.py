"""
8 篇政策深度解讀 IG 發文 + 留言策略試算表

輸出：C:\\Users\\a0920\\teddy-website\\output\\2026-05-22-policy-cards\\policy_posting_plan.xlsx
"""
import sys
if sys.stdout.encoding != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path(r"C:\Users\a0920\teddy-website\output\2026-05-22-policy-cards")
XLSX = OUTPUT_DIR / "policy_posting_plan_v4.xlsx"

# ============ 樣式 ============
HEADER_FILL = PatternFill("solid", fgColor="E27F2E")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12, name="Microsoft JhengHei")
CELL_FONT = Font(size=11, name="Microsoft JhengHei")
WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALT_FILL = PatternFill("solid", fgColor="FFF8F0")


def apply_header(ws, row, columns):
    for col_idx, value in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def apply_row(ws, row_idx, values, alt=False):
    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = CELL_FONT
        cell.alignment = WRAP
        cell.border = BORDER
        if alt:
            cell.fill = ALT_FILL


def auto_height(ws, row_idx, lines):
    """根據文字行數設定 row height"""
    ws.row_dimensions[row_idx].height = max(40, lines * 18)


# ============ Sheet 1：發文 + 留言策略主表 ============
POSTS = [
    {
        "id": "01",
        "topic": "新青安 2.0",
        "date": "5/23（六）21:00",
        "caption": (
            "最近最多客戶問我同一個問題 —\n"
            "「景泰，新青安 2.0 還會有嗎？我是不是該等？」\n\n"
            "2026/7/31 新青安 1.0 補貼到期。\n"
            "目前官方表態：國土署「到期不再補貼」、卓榮泰「續辦但要適度微調」、財政部「6 月底前拍板」。\n\n"
            "我整理目前公開資訊 + 5 張圖卡解讀，幫你判斷該不該現在進場。\n\n"
            "👉 LINE @sky811117 用 AI 算 30 年總利息差多少"
        ),
        "hashtags": "#新青安 #新青安2 #首購 #房貸 #台中房地產",
        "pinned_comment": (
            "📌 留言區常見問題：\n"
            "Q1：新青安 1.0 有年齡上限嗎？ → 1.0 版未設年齡上限，年滿 18 歲即可（細節以公股行庫核貸為準）\n"
            "Q2：第二戶能用新青安嗎？ → 不行，新青安限首購\n"
            "Q3：6 月底還沒拍板，要等嗎？ → 7/31 前送件鎖 1.0，2.0 細節未定不要賭\n\n"
            "想算你的條件能省多少？LINE 我 → sky811117"
        ),
        "extended_comment": (
            "補充：6 月底 2.0 細則出來我會發第二篇追蹤。\n"
            "想收到追蹤推播 → 加 LINE sky811117 / 追 @nov__817"
        ),
        "threads_sync": (
            "新青安 2.0 6 月底前要拍板。目前已知：\n\n"
            "・國土署：到期不再補貼\n"
            "・卓榮泰：續辦但要適度微調\n"
            "・財政部：6 月底前拍板\n\n"
            "想用 1.0 → 7/31 前送件鎖住補貼。\n"
            "細節在我網站完整版（連結 in bio）"
        ),
        "blog_url": "teddy-website-blog.pages.dev/posts/policy-01-new-housing-loan-2-final",
        "shelf_life": "🔴 強時效 — 7/1 後失效，需重發 v2",
    },
    {
        "id": "02",
        "topic": "房地合一 2.0",
        "date": "5/24（日）21:00",
        "caption": (
            "上個月一位客戶賣自住房，賺 480 萬。\n"
            "她開心換房，結果被國稅局課了 168 萬的房地合一稅。\n\n"
            "她以為「我自己住啊，怎麼還要繳這麼多？」\n"
            "問題是她漏了自住豁免 4 條件中的第 3 條。\n\n"
            "👉 滑完 5 張你會知道 4 個級距 + 4 條件\n"
            "👉 LINE @sky811117 賣房前讓我陪你逐條盤點"
        ),
        "hashtags": "#房地合一 #賣房節稅 #自住豁免 #不動產稅 #台中房地產",
        "pinned_comment": (
            "📌 留言區常見問題：\n"
            "Q1：自住豁免一輩子只能用一次嗎？ → 每 6 年可以用一次\n"
            "Q2：戶籍掛在房子但租出去 → 不算自住\n"
            "Q3：我撐到滿 5 年再賣可以省多少？ → 稅率從 35% 降到 20%\n\n"
            "賣房前先算稅，再決定何時賣 → LINE sky811117"
        ),
        "extended_comment": (
            "🆘 真實案例補充：\n"
            "去年成交一筆，賣方差 1 個月過 5 年級距。\n"
            "勸他等 1 個月，稅差 90 萬。最終買賣雙方都贏。\n\n"
            "賣方時機很重要 → LINE 我"
        ),
        "threads_sync": (
            "房地合一 2.0 你以為自住免稅就沒事？\n"
            "4 條件全中才行：\n"
            "✓ 戶籍\n"
            "✓ 持有 + 居住 6 年\n"
            "✓ 6 年內未用過豁免\n"
            "✓ 無營業、無出租\n\n"
            "漏一條都不行 → 完整 IG"
        ),
        "blog_url": "teddy-website-blog.pages.dev/posts/policy-02-house-land-tax-2",
        "shelf_life": "🟢 evergreen — 法規穩定可長期使用",
    },
    {
        "id": "03",
        "topic": "央行信用管制",
        "date": "5/25（一）21:00",
        "caption": (
            "央行 2026/3/20 鬆+緊雙向調整：\n\n"
            "✅ 第二戶 5 → 6 成（鬆綁，換屋族解套）\n"
            "❌ 第三戶以上 4 → 3 成（緊縮）\n"
            "❌ 公司戶 / 高價住宅 4 → 3 成（緊縮）\n\n"
            "換屋族換大房（買 1,800 萬第二戶），自備款從 360 萬變 720 萬 — 多生 360 萬。\n\n"
            "👉 LINE @sky811117 簽約前先預審你能貸幾成"
        ),
        "hashtags": "#央行信用管制 #第二戶 #房貸成數 #換屋族 #台中房地產",
        "pinned_comment": (
            "📌 留言區常見問題：\n"
            "Q1：兩夫妻分頭買可以避開嗎？ → 銀行查家戶，多數會被認定第二戶\n"
            "Q2：先賣舊房再買新房？ → 同日交屋可恢復第一戶資格\n"
            "Q3：公司戶買 → 只能 3 成貸 + 高利率，不建議\n\n"
            "簽約前預審你的成數 → LINE sky811117"
        ),
        "extended_comment": (
            "🆘 簽約前必做：\n"
            "1. 算清楚自己是第幾戶（家戶歸戶）\n"
            "2. 銀行端先預審拿核貸通知書\n"
            "3. 合約寫「貸款不足條款」保護自己"
        ),
        "threads_sync": (
            "央行第 7 波信用管制 = 史上最嚴。\n"
            "第二戶 6 成、第三戶 5.5 成、公司戶 3 成。\n\n"
            "換屋族原本貸 8 成 → 變 6 成\n"
            "= 多自備 360 萬\n\n"
            "對你不是新聞，是真的會卡死你 → IG 完整版"
        ),
        "blog_url": "teddy-website-blog.pages.dev/posts/policy-03-central-bank-credit-control",
        "shelf_life": "🟡 中時效 — 央行 6 月 / 12 月理監事可能調整，每半年 review",
    },
    {
        "id": "04",
        "topic": "不動產說明書",
        "date": "5/26（二）21:00",
        "caption": (
            "去年新聞上一個案子 —\n"
            "屋主賣 1,500 萬的房，沒勾「凶宅」欄位（他以為 20 年前不算）。\n"
            "買方交屋後查到，告上法院，最後判賠 200 萬。\n\n"
            "不動產說明書 30 個欄位，漏勾一個就可能賠錢。\n\n"
            "👉 滑完 5 張看 5 大最容易出事的欄位\n"
            "👉 LINE @sky811117 索取完整 30 欄揭露 checklist"
        ),
        "hashtags": "#不動產說明書 #賣房 #凶宅 #房屋揭露 #房仲教學",
        "pinned_comment": (
            "📌 留言區常見問題：\n"
            "Q1：20 年前的凶宅算嗎？ → 持有期間發生的才需揭，但保險建議全揭\n"
            "Q2：漏水我以為修好了 → 也要揭「曾漏水，已修繕」\n"
            "Q3：頂樓加蓋一定要拆嗎？ → 84/1 前緩拆、之後隨時可拆\n\n"
            "30 欄電子版 checklist → LINE sky811117"
        ),
        "extended_comment": (
            "💡 揭露的賠錢是「減價」\n"
            "💀 不揭露的賠錢是「解約 + 賠償」\n\n"
            "差很多。賣房前先盤一次。"
        ),
        "threads_sync": (
            "不動產說明書 30 欄，賣方最常漏勾 5 個：\n"
            "1. 凶宅（不分年限）\n"
            "2. 漏水（含已修繕）\n"
            "3. 海砂屋（強制驗）\n"
            "4. 違建（84/1 後必揭）\n"
            "5. 嫌惡設施（300m 內）\n\n"
            "「我不知道」不是免責 → IG 詳解"
        ),
        "blog_url": "teddy-website-blog.pages.dev/posts/policy-04-disclosure-document",
        "shelf_life": "🟢 evergreen — 賣方教學常青題材",
    },
    {
        "id": "05",
        "topic": "履約保證",
        "date": "5/27（三）21:00",
        "caption": (
            "去年我簽 1,500 萬的案子，買方看到合約寫「履保費 0.06%」覺得貴。\n"
            "我擋下他想省 4,500 元的念頭。\n\n"
            "3 個月後他朋友買房沒辦履保，賣方拿到頭期款後跑路 ——\n"
            "1,200 萬要不回來。\n\n"
            "👉 滑完 5 張看履保 5 個保護你的情境\n"
            "👉 LINE @sky811117 索取履保合約 checklist"
        ),
        "hashtags": "#履約保證 #買房安全 #建經公司 #信託 #台中房地產",
        "pinned_comment": (
            "📌 留言區常見問題：\n"
            "Q1：哪家建經公司比較好？ → 第一、安信、永信（市占前三）\n"
            "Q2：私下交易也能辦履保嗎？ → 可以，自己找建經公司\n"
            "Q3：履保跟價金信託一樣嗎？ → 不一樣，後者是預售屋\n\n"
            "履保合約眉角 → LINE sky811117"
        ),
        "extended_comment": (
            "💰 履保費 = 保險，不是消費\n"
            "1,500 萬交易 → 你付 4,500\n"
            "保的是 1,500 萬的安全。\n\n"
            "不要省這 4,500。"
        ),
        "threads_sync": (
            "1,500 萬買房，履保費 = 4,500 元。\n"
            "0.06%。\n\n"
            "省了 4,500 → 你冒著 1,500 萬的風險。\n\n"
            "真的不要省。"
        ),
        "blog_url": "teddy-website-blog.pages.dev/posts/policy-05-escrow-system",
        "shelf_life": "🟢 evergreen — 制度穩定，買房通用知識",
    },
    {
        "id": "06",
        "topic": "房屋稅 2.0",
        "date": "5/28（四）21:00",
        "caption": (
            "2025 年 5 月房屋稅開徵，全台多 17 萬戶被多收稅。\n\n"
            "因為囤房稅 2.0 把計算方式從「縣市歸戶」改成「全國歸戶」——\n"
            "你台中 1 間、台北 1 間、高雄 1 間，現在全部加總算。\n\n"
            "👉 滑完 5 張看 4 種自住認定 + 雷區\n"
            "👉 LINE @sky811117 盤點你的全國戶數 + 算稅"
        ),
        "hashtags": "#房屋稅 #囤房稅 #自住認定 #多屋族 #台中房地產",
        "pinned_comment": (
            "📌 留言區常見問題：\n"
            "Q1：全國只有 1 間自住，稅率是？ → 1%（降稅）\n"
            "Q2：我有 4 間，稅金會漲多少？ → 可能漲 50-150%\n"
            "Q3：登記公司行號在家算營業嗎？ → 算，部分喪失自住認定\n\n"
            "盤點你的全國戶數 → LINE sky811117"
        ),
        "extended_comment": (
            "📊 房屋稅 2.0 + 房地合一 2.0 + 央行管制\n"
            "= 三重打擊，多屋族要算總帳。\n\n"
            "別只看單一稅，要看 5 年總成本。"
        ),
        "threads_sync": (
            "房屋稅 2.0 上路後 ——\n"
            "全國 1 間自住 → 1%（降）\n"
            "4 間以上 → 2.0-4.8%（漲）\n\n"
            "全台 17 萬戶被多收稅。\n"
            "因為「全國歸戶」啟動了。"
        ),
        "blog_url": "teddy-website-blog.pages.dev/posts/policy-06-house-tax-2",
        "shelf_life": "🟡 中時效 — 4 月底 / 5 月初開徵前再發一次（提醒申辦）",
    },
    {
        "id": "07",
        "topic": "地價稅",
        "date": "5/29（五）21:00",
        "caption": (
            "去年 11 月地價稅開徵，客戶林姐拿稅單來找我 ——\n"
            "「景泰，我這張 4 萬 8，鄰居一樣坪數只繳 1 萬 2。差 4 倍。」\n\n"
            "查到她沒申辦自用住宅優惠稅率。\n"
            "0.2% vs 1.0% = 差 5 倍。8 年累積差 28 萬。\n\n"
            "👉 滑完 5 張看自用住宅 4 條件\n"
            "👉 9/22 前申辦，當年才適用"
        ),
        "hashtags": "#地價稅 #自用住宅 #持有稅 #公告地價 #節稅",
        "pinned_comment": (
            "📌 留言區常見問題：\n"
            "Q1：什麼時候辦自用住宅？ → 9/22 前完成，當年適用\n"
            "Q2：跨縣市分散有用嗎？ → 全國 3 處內歸戶，分散沒差\n"
            "Q3：透天厝為什麼地價稅特別重？ → 土地持分 100%\n\n"
            "申辦流程 + 文件清單 → LINE sky811117"
        ),
        "extended_comment": (
            "⏰ 9/22 前申辦自用住宅優惠\n"
            "錯過要等明年。\n"
            "每年省 1-7 萬，8 年就是 1 台車的錢。"
        ),
        "threads_sync": (
            "地價稅不是「自動」給你優惠。\n"
            "是「自己申辦」才有。\n\n"
            "自用住宅 0.2% vs 一般 1.0-5.5%\n"
            "差 5-27 倍。\n\n"
            "9/22 前申辦 → 當年才適用。"
        ),
        "blog_url": "teddy-website-blog.pages.dev/posts/policy-07-land-value-tax",
        "shelf_life": "🟡 中時效 — 9 月初再發一次（搶 9/22 截止）+ 11 月開徵前",
    },
    {
        "id": "08",
        "topic": "平均地權條例",
        "date": "5/30（六）21:00",
        "caption": (
            "2023/7/1 平均地權條例修正 ——\n\n"
            "・預售屋換約轉售 → 限親屬/非自願/同戶換\n"
            "・建商協助轉售 → 每戶罰 50-300 萬\n"
            "・私法人購住宅 → 須許可、5 年不得移轉\n"
            "・散布不實炒作 → 罰 100-5,000 萬\n"
            "・檢舉獎金 → 罰金 30%、單案上限 1,000 萬\n\n"
            "預售屋現在是「買來住的、不是買來炒的」。\n\n"
            "👉 LINE @sky811117 預售屋規劃前讓我幫你評估"
        ),
        "hashtags": "#平均地權條例 #預售屋 #紅單禁炒 #私法人購屋 #台中房地產",
        "pinned_comment": (
            "📌 留言區常見問題：\n"
            "Q1：預售屋買來不住能轉嗎？ → 不能（除親屬 / 非自願 / 同戶換 / 內政部公告之 6 種特殊情形）\n"
            "Q2：公司戶可以買住宅嗎？ → 要申請許可，5 年內不得移轉\n"
            "Q3：紅單轉手會被抓嗎？ → 檢舉獎金 30%、上限 1,000 萬，同業 / 員工 / 客戶都可舉發\n\n"
            "預售屋規劃 → LINE sky811117"
        ),
        "extended_comment": (
            "🚨 預售屋現在是「買來住的、不是買來炒的」\n"
            "規則改了，玩法也要改。\n\n"
            "投資族該轉向：中古屋 + 商辦 + 危老都更。"
        ),
        "threads_sync": (
            "預售屋現在是「買來住的、不是買來炒的」。\n\n"
            "建商協助轉售 → 50-300 萬/戶\n"
            "散布不實炒作 → 100-5,000 萬\n"
            "檢舉獎金 → 30%、上限 1,000 萬\n\n"
            "規則改了，玩法也要改。"
        ),
        "blog_url": "teddy-website-blog.pages.dev/posts/policy-08-presale-equality-amendment",
        "shelf_life": "🟢 evergreen — 修法已落地，長期教材",
    },
]


def build_main_sheet(wb):
    ws = wb.active
    ws.title = "發文留言策略"

    title = ws.cell(row=1, column=1, value="陳景泰 · 8 篇政策深度解讀 IG 發文 + 留言策略")
    title.font = Font(bold=True, size=18, color="E27F2E", name="Microsoft JhengHei")
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 36

    note = ws.cell(row=2, column=1, value="每篇 5 張圖卡（cover / bullet1 / number / bullet2 / cta）；發文後立刻在自己貼文留 pin 留言（引導 LINE）+ 5 分鐘後再留 extended 補充。Threads / FB 同步在 caption 改寫版。")
    note.font = Font(size=10, color="888888", italic=True, name="Microsoft JhengHei")
    note.alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 38

    columns = [
        "篇號",
        "主題",
        "發文時間",
        "IG Caption（≤150 字）",
        "5 個 Hashtag",
        "📌 Pin 留言（發文後立刻 / 自留）",
        "💬 延伸留言（5 分鐘後 / 自留）",
        "Threads / FB 同步版",
        "🔗 Blog URL（完整版連結）",
        "⏱ 保留期 / 重發提醒",
    ]
    apply_header(ws, 4, columns)
    ws.row_dimensions[4].height = 30

    for i, post in enumerate(POSTS):
        row = 5 + i
        values = [
            post["id"],
            post["topic"],
            post["date"],
            post["caption"],
            post["hashtags"],
            post["pinned_comment"],
            post["extended_comment"],
            post["threads_sync"],
            post["blog_url"],
            post["shelf_life"],
        ]
        apply_row(ws, row, values, alt=(i % 2 == 1))
        # 自動 height：最長那一欄的行數
        max_lines = max((str(v).count("\n") + 1) for v in values)
        ws.row_dimensions[row].height = max_lines * 18 + 10

    # Column widths
    widths = [6, 16, 16, 50, 35, 50, 35, 40, 42, 32]
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


# ============ Sheet 2：FAQ 標準回應 ============
FAQ = [
    {
        "category": "新青安",
        "q": "我超過 40 歲還能用新青安嗎？",
        "a": "目前 1.0 限 40 歲以下。2.0（6 月底定案）可能放寬到 45 歲。如果你 40-45 歲，建議 7 月初再看細則決定。要算月付能力可以 LINE 我 sky811117。",
    },
    {
        "category": "新青安",
        "q": "新青安 vs 一般房貸差多少？",
        "a": "借 800 萬 30 年，新青安 1.0 約省 35 萬利息（0.375% 補貼）。2.0 補貼可能降到 0.25% → 省 23 萬。可以 LINE 我用 AI 算你的條件。",
    },
    {
        "category": "房地合一",
        "q": "自住豁免一輩子只能用一次嗎？",
        "a": "不是。是「6 年間隔 1 次」—— 上次用過後 6 年才能再用第二次。可以用很多次，但每兩次間隔 6 年。",
    },
    {
        "category": "房地合一",
        "q": "我撐到 5 年再賣可以省多少？",
        "a": "稅率從 35% 降到 20%。賺 500 萬 → 稅從 175 萬降到 100 萬，省 75 萬。差 1 天差很大，可以 LINE 我幫你算。",
    },
    {
        "category": "央行管制",
        "q": "兩夫妻分開買，可以避開第二戶嗎？",
        "a": "理論上可以、實務上難。銀行會查家戶狀況：戶籍、共同生活、家庭已有自住房 → 通常認定第二戶。完全分居才可能例外。",
    },
    {
        "category": "央行管制",
        "q": "換屋族怎麼解？",
        "a": "3 個解法：1) 先賣舊房再買新房；2) 同日交屋；3) 二胎 / 信貸補頭款（不建議）。LINE 我評估你的最佳路徑。",
    },
    {
        "category": "不動產說明書",
        "q": "20 年前的凶宅算嗎？",
        "a": "法規上「賣方持有期間發生」才需揭。20 年前是前手不是你 → 法規上可不揭。但保險建議全揭，買方查到還是可能告。",
    },
    {
        "category": "不動產說明書",
        "q": "我以為漏水修好了還要寫嗎？",
        "a": "要。建議寫「99 年 X 月曾漏水，已修繕」。修了還是要揭，因為買方有權知道歷史狀況。",
    },
    {
        "category": "履約保證",
        "q": "履保費誰付？",
        "a": "買賣雙方各半。1,500 萬交易 → 各付 4,500 元。真的不要省這 4,500，保的是 1,500 萬安全。",
    },
    {
        "category": "履約保證",
        "q": "私下交易也能辦履保嗎？",
        "a": "可以。不透過仲介自己直接找建經公司（第一、安信、永信）辦履保。費率還是 0.06%。",
    },
    {
        "category": "房屋稅",
        "q": "我自住卻被課 1.2% 不是 1%？",
        "a": "「全國單一自住」才有 1%。如果你全家全國有 2 間以上 → 1.2%。確認家戶名下幾間。",
    },
    {
        "category": "房屋稅",
        "q": "登記公司在家算營業嗎？",
        "a": "算。即便沒實際營業，登記就會被認定。建議單純自住的房不要登記公司行號。",
    },
    {
        "category": "地價稅",
        "q": "什麼時候要申辦自用住宅？",
        "a": "每年 9/22 前。錯過要等明年。每年省 1-7 萬不等，累積很可觀。LINE 我可以給你流程 + 文件清單。",
    },
    {
        "category": "地價稅",
        "q": "透天厝為什麼地價稅特別重？",
        "a": "因為土地持分 100%。大樓住戶土地持分只有 1-3 坪，透天可能 20-50 坪。同樣地段地價稅可能差 10 倍以上。",
    },
    {
        "category": "平均地權",
        "q": "預售屋買來不住能轉嗎？",
        "a": "不能（除：親屬間、非自願失業 / 重大傷病 / 災害、同建案換戶）。違法每戶罰 15-100 萬。",
    },
    {
        "category": "平均地權",
        "q": "我朋友說可以轉，是真的嗎？",
        "a": "2023/7/1 後不行。檢舉獎金 20%（單案上限 1,000 萬）—— 同業 / 客戶 / 員工都可舉發。風險太大。",
    },
]


def build_faq_sheet(wb):
    ws = wb.create_sheet("FAQ 標準回應")

    title = ws.cell(row=1, column=1, value="客戶常見問題 + 標準回應（複製貼回留言即可）")
    title.font = Font(bold=True, size=16, color="E27F2E", name="Microsoft JhengHei")
    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 30

    apply_header(ws, 3, ["分類", "客戶可能問", "標準回應"])
    ws.row_dimensions[3].height = 28

    for i, item in enumerate(FAQ):
        row = 4 + i
        apply_row(ws, row, [item["category"], item["q"], item["a"]], alt=(i % 2 == 1))
        ws.row_dimensions[row].height = max(50, item["a"].count("\n") * 18 + 50)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 65


# ============ Sheet 3：KPI 追蹤 ============
def build_kpi_sheet(wb):
    ws = wb.create_sheet("KPI 追蹤")

    title = ws.cell(row=1, column=1, value="8 篇政策文 IG 表現追蹤（發文後 7 天填一次）")
    title.font = Font(bold=True, size=16, color="E27F2E", name="Microsoft JhengHei")
    ws.merge_cells("A1:I1")
    ws.row_dimensions[1].height = 30

    note = ws.cell(row=2, column=1, value="每篇貼文發出後第 7 天回來填數據。觀察哪個主題最會帶來「實際 LINE 加好友」（這才是真的 ROI）。")
    note.font = Font(size=10, color="888888", italic=True, name="Microsoft JhengHei")
    note.alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:I2")
    ws.row_dimensions[2].height = 30

    columns = [
        "篇號", "主題", "發文日", "讚數", "留言數", "收藏數", "分享數", "私訊數", "LINE 加好友（ROI）"
    ]
    apply_header(ws, 4, columns)
    ws.row_dimensions[4].height = 28

    for i, post in enumerate(POSTS):
        row = 5 + i
        values = [post["id"], post["topic"], post["date"].split("（")[0], "", "", "", "", "", ""]
        apply_row(ws, row, values, alt=(i % 2 == 1))
        ws.row_dimensions[row].height = 30

    # 總計列
    total_row = 5 + len(POSTS)
    ws.cell(row=total_row, column=1, value="").fill = PatternFill("solid", fgColor="F0F0F0")
    ws.cell(row=total_row, column=2, value="總計").font = Font(bold=True, color="E27F2E", name="Microsoft JhengHei")
    for col_idx in [4, 5, 6, 7, 8, 9]:
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=total_row, column=col_idx, value=f"=SUM({col_letter}5:{col_letter}{total_row-1})")
        cell.font = Font(bold=True, name="Microsoft JhengHei")
        cell.alignment = CENTER
        cell.border = BORDER
        cell.fill = PatternFill("solid", fgColor="FFF0E0")

    widths = [6, 18, 14, 10, 10, 10, 10, 10, 18]
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


# ============ Sheet 4：發文 Checklist ============
def build_checklist_sheet(wb):
    ws = wb.create_sheet("發文 Checklist")

    title = ws.cell(row=1, column=1, value="每篇發文 Checklist（從畫圖卡到推播全流程）")
    title.font = Font(bold=True, size=16, color="E27F2E", name="Microsoft JhengHei")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 30

    items = [
        ("T-1 天", "5 張圖卡 review 是否要修文字 / 數據"),
        ("T-1 天", "Caption 從試算表複製、檢查 5 個 hashtag 是否到位"),
        ("發文前 1 小時", "確認 21:00 排程或手動準備"),
        ("21:00 發文", "IG 輪播 5 張、selected hashtag 用單獨留言（不放 caption）"),
        ("21:00", "Threads / FB 同步發（改寫版）"),
        ("21:01 立刻", "自己留 📌 Pin 留言（FAQ 摘要 + LINE 引導）"),
        ("21:01", "Pin 自己這條留言（IG 點 ... → 釘選）"),
        ("21:05", "自己留 💬 延伸留言（補充 1 句金句）"),
        ("21:30", "Story 轉貼（限動 9:16）"),
        ("22:00", "LINE 群推連結 + 「我剛發了 IG 政策解讀」+ Blog 全文連結"),
        ("發文後 24 小時", "回覆所有留言（黑名單同行排除）"),
        ("發文後 7 天", "回填 KPI Sheet 數據"),
        ("發文後 7 天", "私訊主動關心 → 高互動但沒加 LINE 的觀眾"),
    ]

    apply_header(ws, 3, ["時間點", "動作"])
    ws.row_dimensions[3].height = 28

    for i, (when, what) in enumerate(items):
        row = 4 + i
        apply_row(ws, row, [when, what], alt=(i % 2 == 1))
        ws.row_dimensions[row].height = 30

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 65


# ============ Sheet 5：留言應對策略（4 類型） ============
COMMENT_STRATEGY = [
    {
        "type": "🟢 真客戶提問",
        "signal": (
            "• 有具體個人狀況（「我 35 歲想買北屯...」）\n"
            "• 提到家人 / 預算 / 區域\n"
            "• 留言 + 表情很真誠\n"
            "• 帳號活躍、有真人感"
        ),
        "response": (
            "1️⃣ 5 分鐘內回應（黃金時段）\n"
            "2️⃣ 公開留言：簡答 + 「私訊更詳細」\n"
            "3️⃣ 主動 DM：個人化稱呼 + 詢問補充資訊\n"
            "4️⃣ 引導加 LINE：「我把試算表寄給你」\n"
            "5️⃣ Notion 加客戶名單（標：來自 IG / 政策文 / X 篇）"
        ),
        "example": (
            "客：「我 32 歲，想買北屯 1500 萬的房，新青安能用嗎？」\n\n"
            "回：「能用！32 歲首購正中央。但要看你預算配比 — 私訊我你的年收與自備款，我用 AI 算 30 年總利息給你看，5 分鐘的事。」"
        ),
    },
    {
        "type": "🔵 純讚 / 短評",
        "signal": (
            "• 「推」「實用」「好文」\n"
            "• 表情符號為主\n"
            "• 沒提問題、沒個人狀況"
        ),
        "response": (
            "1️⃣ 1 小時內愛心 + 短回\n"
            "2️⃣ 不引導 LINE（會 spam 感）\n"
            "3️⃣ 反問一個輕問題開話匣\n"
            "4️⃣ 追蹤對方 IG 看是否潛在客"
        ),
        "example": (
            "客：「謝謝整理 👍」\n\n"
            "回：「不客氣！這 8 篇是我這週連發的政策系列，明天會出房地合一 2.0 的自住豁免 4 條件 — 你目前有自住房嗎？」"
        ),
    },
    {
        "type": "🟡 同行 / 撈客戶（黑名單候選）",
        "signal": (
            "• IG handle 有「房仲」「不動產」「永慶」「信義」「住商」\n"
            "• 留言「歡迎找我詳談」「我也可以幫你」\n"
            "• 私訊客戶疑似撈客\n"
            "• 帳號 bio 同行品牌"
        ),
        "response": (
            "1️⃣ 公開留言不回（不要互動拉高他曝光）\n"
            "2️⃣ 紀錄到「同行黑名單」（下方欄位）\n"
            "3️⃣ 必要時用 IG 限制功能\n"
            "4️⃣ 同時 BigKanBan boss-comment-monitor 對該 handle 加 ignore_handles"
        ),
        "example": (
            "同行：「房地合一的部分我這邊也很熟，可以找我！」\n\n"
            "回應方式：❌ 不要回，會讓他在我的曝光下撈客。\n"
            "✅ 紀錄黑名單，下次他來留言不再點開、不互動。"
        ),
    },
    {
        "type": "🔴 質疑 / 挑釁 / 酸民",
        "signal": (
            "• 「房仲都騙人」「政府的話也信」\n"
            "• 抓字面意思過度解讀\n"
            "• 明顯不買房只是發洩"
        ),
        "response": (
            "1️⃣ 不被激怒、不辯論細節\n"
            "2️⃣ 公開回：理性、簡短、附法規連結\n"
            "3️⃣ 不刪除（會被截圖傳）\n"
            "4️⃣ 持續挑釁 → 隱藏留言（不刪）\n"
            "5️⃣ 真實人身攻擊 → 檢舉 + 封鎖"
        ),
        "example": (
            "酸：「自住豁免條件這麼多，根本是政府詐財」\n\n"
            "回：「條件確實多，所以我才寫這篇幫大家對。法規依據是房地合一稅 §4-5 第 2 項，國稅局網站有完整版。」（附連結，不再回他第二輪）"
        ),
    },
]


def build_comment_strategy_sheet(wb):
    ws = wb.create_sheet("留言應對策略")

    title = ws.cell(row=1, column=1, value="4 類留言 × 應對策略（黃金回應時間：發文後 1 小時內）")
    title.font = Font(bold=True, size=16, color="E27F2E", name="Microsoft JhengHei")
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 30

    note = ws.cell(row=2, column=1, value="留言區戰略「比同行快」：客戶提問 5 分鐘內回 + 引導 LINE；同行撈客不要互動；酸民冷處理不辯論。BigKanBan boss-comment-monitor 的同行黑名單跨用。")
    note.font = Font(size=10, color="888888", italic=True, name="Microsoft JhengHei")
    note.alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:D2")
    ws.row_dimensions[2].height = 36

    apply_header(ws, 4, ["留言類型", "辨識信號", "應對 SOP（5 步）", "實例（情境 + 標準回應）"])
    ws.row_dimensions[4].height = 30

    for i, item in enumerate(COMMENT_STRATEGY):
        row = 5 + i
        values = [item["type"], item["signal"], item["response"], item["example"]]
        apply_row(ws, row, values, alt=(i % 2 == 1))
        max_lines = max(str(v).count("\n") + 1 for v in values)
        ws.row_dimensions[row].height = max_lines * 18 + 10

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 55

    # 同行黑名單區
    bl_start = 5 + len(COMMENT_STRATEGY) + 2
    bl_title = ws.cell(row=bl_start, column=1, value="📋 同行黑名單（IG handle 紀錄，每次遇到就加，跨用到 boss-comment-monitor）")
    bl_title.font = Font(bold=True, size=14, color="E27F2E", name="Microsoft JhengHei")
    ws.merge_cells(f"A{bl_start}:D{bl_start}")
    ws.row_dimensions[bl_start].height = 26

    apply_header(ws, bl_start + 1, ["日期", "IG Handle", "公司 / 品牌", "備註"])
    # 留 15 列空白
    for i in range(15):
        row = bl_start + 2 + i
        apply_row(ws, row, ["", "", "", ""], alt=(i % 2 == 1))
        ws.row_dimensions[row].height = 26


# ============ Sheet 6：跨平台分發 tracker ============
def build_distribution_sheet(wb):
    ws = wb.create_sheet("跨平台分發 tracker")

    title = ws.cell(row=1, column=1, value="8 篇 × 5 平台分發狀態（發完就打✓）")
    title.font = Font(bold=True, size=16, color="E27F2E", name="Microsoft JhengHei")
    ws.merge_cells("A1:I1")
    ws.row_dimensions[1].height = 30

    note = ws.cell(row=2, column=1, value="一篇文發 5 處 = 完整觸點。Story / LINE 群是 dark social（看不到統計但點擊率高）。")
    note.font = Font(size=10, color="888888", italic=True, name="Microsoft JhengHei")
    note.alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:I2")
    ws.row_dimensions[2].height = 26

    columns = [
        "篇號", "主題", "發文日",
        "IG 輪播", "IG Story", "Threads", "FB 粉專", "LINE 群推", "Blog 上線（draft→false）"
    ]
    apply_header(ws, 4, columns)
    ws.row_dimensions[4].height = 30

    for i, post in enumerate(POSTS):
        row = 5 + i
        values = [
            post["id"],
            post["topic"],
            post["date"].split("（")[0],
            "☐", "☐", "☐", "☐", "☐", "☐",
        ]
        apply_row(ws, row, values, alt=(i % 2 == 1))
        for col_idx in [4, 5, 6, 7, 8, 9]:
            ws.cell(row=row, column=col_idx).alignment = CENTER
        ws.row_dimensions[row].height = 32

    widths = [6, 18, 14, 12, 12, 12, 12, 12, 18]
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


# ============ 主程式 ============

def main():
    wb = Workbook()
    build_main_sheet(wb)
    build_faq_sheet(wb)
    build_comment_strategy_sheet(wb)
    build_distribution_sheet(wb)
    build_kpi_sheet(wb)
    build_checklist_sheet(wb)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX)
    print(f"✅ 已產生：{XLSX}")
    print(f"   - Sheet 1：發文留言策略（8 列 × 10 欄）")
    print(f"   - Sheet 2：FAQ 標準回應（{len(FAQ)} 列）")
    print(f"   - Sheet 3：留言應對策略（4 類型 + 15 列黑名單空白）")
    print(f"   - Sheet 4：跨平台分發 tracker（8 × 5 平台 checkbox）")
    print(f"   - Sheet 5：KPI 追蹤（自填）")
    print(f"   - Sheet 6：發文 Checklist（13 步）")


if __name__ == "__main__":
    main()
